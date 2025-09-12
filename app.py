from flask import Flask, render_template, request, redirect, url_for, session, flash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, json

app = Flask(__name__)
app.secret_key = 'adm12345*' # Key de Flask, no es algo importante 
 
PRODUCTS_FILE = 'products.xlsx' # Definimos PRODUCTS_FILE Y ORDERS_FILE para poder llamar los documentos luego en el codigo
ORDERS_FILE = 'orders.xlsx'

def ensure_files(): # Se verifica si estan los archivos excel donde vamos a guardar los productos y los pedidos respectivamente, si no existen, se crean con los nombres de las columnas que vamos a necesitar.
    if not os.path.exists(PRODUCTS_FILE): 
        wb = Workbook(); ws = wb.active; ws.title = 'products' 
        ws.append(['id','name','model','color','size','price','stock']); wb.save(PRODUCTS_FILE)
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook(); ws = wb.active; ws.title = 'orders' 
        ws.append(['id','customer_name','address','phone','deadline','items_json','total_price']); wb.save(ORDERS_FILE) # Luego de agregarlas, se guardan 
 
def next_id(path, sheet):
    wb = load_workbook(path); ws = wb[sheet]
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            try:
                v = int(row[0])
                if v > max_id: max_id = v
            except: pass
    wb.close(); return max_id + 1

def load_products(): # Función para cargar los productos registrados
    wb = load_workbook(PRODUCTS_FILE); ws = wb['products']; products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        products.append({'id':int(row[0]), 'name':row[1], 'model':row[2], 'color':row[3], 'size':row[4], 'price':float(row[5] or 0), 'stock':int(row[6] or 0)})
    wb.close(); return products

def save_product(prod): # Función para guardar un producto (argumento peoducto)
    wb = load_workbook(PRODUCTS_FILE); ws = wb['products']; found=False
    for r in ws.iter_rows(min_row=2):
        if r[0].value == prod.get('id'):
            r[1].value = prod.get('name'); r[2].value = prod.get('model'); r[3].value = prod.get('color')
            r[4].value = prod.get('size'); r[5].value = prod.get('price'); r[6].value = prod.get('stock')
            found = True; break
    if not found:
        nid = prod.get('id') or next_id(PRODUCTS_FILE, 'products'); ws.append([nid, prod.get('name'), prod.get('model'), prod.get('color'), prod.get('size'), prod.get('price'), prod.get('stock')])
    wb.save(PRODUCTS_FILE); wb.close()

def delete_product(pid): # Funcion para eliminar un producto 
    wb = load_workbook(PRODUCTS_FILE); ws = wb['products']; row_to_delete=None
    for idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
        if r[0].value == pid: row_to_delete = idx; break
    if row_to_delete: ws.delete_rows(row_to_delete) # Elimina toda la fila 
    wb.save(PRODUCTS_FILE); wb.close()

def append_order(order): # Con esto se agregan las órdenes a la lista de pedidos
    wb = load_workbook(ORDERS_FILE); ws = wb['orders']; nid = next_id(ORDERS_FILE, 'orders')
    ws.append([nid, order.get('customer_name'), order.get('address'), order.get('phone'), order.get('deadline'), order.get('items_json'), order.get('total_price')])
    wb.save(ORDERS_FILE); wb.close()


def delete_order(oid): # Vamos a eliminar la orden que queramos cuando se marque como completa
    wb = load_workbook(ORDERS_FILE)
    ws = wb['orders']
    row_to_delete = None
    for idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
        if r[0].value == oid:
            row_to_delete = idx
            break
    if row_to_delete:
        ws.delete_rows(row_to_delete)
    wb.save(ORDERS_FILE)
    wb.close()

def load_orders(): # Se cargan los pedidos que ya están guardados 
    wb = load_workbook(ORDERS_FILE); ws = wb['orders']; orders=[]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        orders.append({'id':row[0], 'customer_name':row[1], 'address':row[2], 'phone':row[3], 'deadline':row[4], 'items':json.loads(row[5] or '[]'), 'total_price':float(row[6] or 0)})
    wb.close(); return orders

def is_logged_in(): # Se asigna como valor predeterminado al entrar a la página que no tiene sesión iniciada
    return session.get('logged_in', False)

def format_currency_colombian(amount): # Función para formatear precios en pesos colombianos
    return f"${amount:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

@app.route('/login', methods=['GET','POST'])
def login():
    if is_logged_in(): return redirect(url_for('products')) # Si ya tiene la sesión iniciada, redirige a productos
    error = None
    if request.method == 'POST':
        u = request.form.get('username'); p = request.form.get('password') # Se obtiene del formulario del login el usuario y la contraseña
        if u == 'admin' and p == 'admin': # Definimos con qué se puede entrar a la pagina, si no es con ese usuario y contraseña, no entra.
            session['logged_in'] = True
            return redirect(url_for('products'))
        else:
            error = 'Credenciales inválidas (usa el usuario y contraseña que se encuentra en readme'
    return render_template('login.html', error=error) 

@app.route('/logout') # Esto es para que cuando le demos a logout o cerrar sesión, cierre la sesión y nos lleve a login
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/') # Funcion para cuando se quiere redirigir a una parte de la pagina.
def root():
    if not is_logged_in(): return redirect(url_for('login'))
    return redirect(url_for('products'))

@app.route('/products') # Con esto cargamos los productos y los mostramos
def products():
    if not is_logged_in(): return redirect(url_for('login'))
    products = load_products() # Llamamos la funcion de cargar productos como products para usarlo luego.
    return render_template('products.html', products=products) 

@app.route('/product/new', methods=['GET','POST']) # Esta función nos permite crear los productos, solicita nombre, modelo, color, talla, precio, stock, respectivamente, luego lo guarda como prod junto a la hora de creación 
def product_new():
    if not is_logged_in(): return redirect(url_for('login'))
    if request.method == 'POST':
        name = request.form.get('name'); model = request.form.get('model'); color = request.form.get('color')
        size = request.form.get('size'); price = float(request.form.get('price') or 0); stock = int(request.form.get('stock') or 0)
        prod = {'name': name, 'model': model, 'color': color, 'size': size, 'price': price, 'stock': stock} # Definimos cada variable solicitandolo en el formulario del nuevo producto para luego guardarlo en la base de datos.
        save_product(prod); return redirect(url_for('products'))
    return render_template('product_form.html', product=None)

@app.route('/product/edit/<int:pid>', methods=['GET','POST'])  
def product_edit(pid): # Si queremos editar un producto, elegimos id del producto y esto nos lleva a la página donde lo podemos modificar 
    if not is_logged_in(): return redirect(url_for('login'))
    products = load_products(); p = next((x for x in products if x['id']==pid), None)
    if not p: flash('Producto no encontrado'); return redirect(url_for('products')) # Si uno busca un producto que no existe (id), muestra el mensaje
    if request.method == 'POST':
        p['name'] = request.form.get('name'); p['model'] = request.form.get('model'); p['color'] = request.form.get('color')
        p['size'] = request.form.get('size'); p['price'] = float(request.form.get('price') or 0); p['stock'] = int(request.form.get('stock') or 0)
        save_product(p); return redirect(url_for('products'))
    return render_template('product_form.html', product=p)

@app.route('/product/delete/<int:pid>', methods=['POST'])
def product_delete(pid): # Si queremos eliminar un producto con eso lo borramos de la base de datos con el product id
    if not is_logged_in(): return redirect(url_for('login'))
    delete_product(pid); return redirect(url_for('products'))

@app.route('/orders') # Establece para cuando se necesite redirigir a orders 
def orders():
    if not is_logged_in(): return redirect(url_for('login'))
    orders = load_orders(); return render_template('orders.html', orders=orders) # La funcion orders para cargar los pedidos y llevar al html de pedidos.

@app.route('/order/new', methods=['GET','POST'])
def order_new(): # Creamos ordenes, solicita el nombre del cliente, dirección, teléfono y producto con su cantidad y fecha de entrega.
    if not is_logged_in(): return redirect(url_for('login'))
    products = load_products()

    if request.method == 'POST':
        customer_name = request.form.get('customer_name'); address = request.form.get('address'); phone = request.form.get('phone'); deadline = request.form.get('deadline')
        ids = request.form.getlist('product_id'); qtys = request.form.getlist('qty')
        items = []; total = 0.0; prod_map = {p['id']:p for p in products} # Acá esta el mapeo con un bucle en products.

        for i,pid_raw in enumerate(ids):
            if not pid_raw: continue
            pid = int(pid_raw); qty = int(qtys[i] or 1) # Definimos pid (id del producto) y Qty (Cantidades)
            prod = prod_map.get(pid) # hace un mapeo (Revisa cada id en la lista de productos) para conseguir el producto con el pid especifico.
            if not prod: continue
            if prod['stock'] < qty:
                flash(f'Stock insuficiente para {prod["name"]}, por favor ingrese un número entre 1 y {prod['stock']}'); return redirect(url_for('order_new')) # Si el stock es menos que la cantidad de elementos que se estan pidiendo, devuelve a generar un nuevo pedido, basicamente lo cancela.

        for i,pid_raw in enumerate(ids):
            if not pid_raw: continue
            pid = int(pid_raw); qty = int(qtys[i] or 1)
            prod = prod_map.get(pid); prod['stock'] -= qty; save_product(prod) # Al crear el pedido se actualiza la cantidad de stock que hay en la base de datos de cada producto solicitado en el pedido y se guarda.
            items.append({'id':pid,'name':prod['name'],'qty':qty,'price':prod['price'],'size': prod.get('size'), 'color': prod.get('color'), 'deadline':prod.get('deadline')}); total += prod['price'] * qty
        order = {'customer_name':customer_name,'address':address,'phone':phone, 'deadline':deadline, 'items_json':json.dumps(items),'total_price':total}
        append_order(order); return redirect(url_for('orders')) # Se guarda el pedido y se lleva a la pagina de pedidos.
    return render_template('order_form.html', products=products)

@app.route('/order/delete/<int:oid>', methods=['POST'])
def order_delete(oid):  # Cuando clickeamos el boton de completar pedido muestra el mensaje de la linea 177, y elimina la orden.
    if not is_logged_in():
        return redirect(url_for('login')) # Al marcar como completado un pedido, este es eliminado de la base de datos y muestra un mensaje cuando lo elimina.
    delete_order(oid)
    flash('Pedido marcado como completado y eliminado')
    return redirect(url_for('orders'))

@app.context_processor
def inject_format_currency(): # Esta funcion aplcia formato de dinero en pesos colombianos a las secciones donde hayan precios.
    return dict(format_currency=format_currency_colombian)

if __name__ == '__main__':
    ensure_files(); app.run(debug=True)
