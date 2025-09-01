from openpyxl import Workbook, load_workbook
from datetime import datetime
import json, os

PRODUCTS_XLSX = 'products.xlsx'
ORDERS_XLSX = 'orders.xlsx'
LOGS_XLSX = 'stock_logs.xlsx'
USERS_JSON = 'users.json'

def ensure_files():
    if not os.path.exists(USERS_JSON):
        from werkzeug.security import generate_password_hash
        admin = {'users': [{'id': 1, 'username': 'admin', 'password_hash': generate_password_hash('admin'), 'is_admin': True}]}
        with open(USERS_JSON, 'w', encoding='utf-8') as f:
            json.dump(admin, f)
    if not os.path.exists(PRODUCTS_XLSX):
        wb = Workbook(); ws = wb.active; ws.title='products'
        ws.append(['id','name','category','model','color','size','price','stock','created_at']); wb.save(PRODUCTS_XLSX)
    if not os.path.exists(ORDERS_XLSX):
        wb = Workbook(); ws = wb.active; ws.title='orders'
        ws.append(['id','customer_name','address','city','phone','items_json','total_price','created_at']); wb.save(ORDERS_XLSX)
    if not os.path.exists(LOGS_XLSX):
        wb = Workbook(); ws = wb.active; ws.title='logs'
        ws.append(['id','product_id','change_amount','reason','username','timestamp']); wb.save(LOGS_XLSX)

def _next_id_from_sheet(path, sheet_name):
    wb = load_workbook(path); ws = wb[sheet_name]
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        try:
            rid = int(row[0])
            if rid > max_id: max_id = rid
        except: pass
    wb.close(); return max_id+1

def load_products():
    wb = load_workbook(PRODUCTS_XLSX); ws = wb['products']; products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        products.append({'id':int(row[0]), 'name':row[1], 'category':row[2], 'model':row[3], 'color':row[4], 'size':row[5], 'price':float(row[6] or 0), 'stock':int(row[7] or 0), 'created_at':row[8]})
    wb.close(); return products

def save_product_row(product):
    wb = load_workbook(PRODUCTS_XLSX); ws = wb['products']; found=False
    for r in ws.iter_rows(min_row=2):
        if r[0].value == product.get('id'):
            r[1].value = product.get('name'); r[2].value = product.get('category'); r[3].value = product.get('model')
            r[4].value = product.get('color'); r[5].value = product.get('size'); r[6].value = product.get('price')
            r[7].value = product.get('stock'); r[8].value = product.get('created_at'); found=True; break
    if not found:
        nid = product.get('id') or _next_id_from_sheet(PRODUCTS_XLSX,'products')
        ws.append([nid, product.get('name'), product.get('category'), product.get('model'), product.get('color'), product.get('size'), product.get('price'), product.get('stock'), product.get('created_at')])
    wb.save(PRODUCTS_XLSX); wb.close()

def delete_product_by_id(pid):
    wb = load_workbook(PRODUCTS_XLSX); ws = wb['products']; row_to_delete=None
    for idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
        if r[0].value == pid: row_to_delete = idx; break
    if row_to_delete: ws.delete_rows(row_to_delete)
    wb.save(PRODUCTS_XLSX); wb.close()

def append_log(product_id, change_amount, reason, username):
    wb = load_workbook(LOGS_XLSX); ws = wb['logs']; nid = _next_id_from_sheet(LOGS_XLSX,'logs')
    ts = datetime.utcnow().isoformat(); ws.append([nid, product_id, change_amount, reason, username, ts]); wb.save(LOGS_XLSX); wb.close()

def append_order(order):
    wb = load_workbook(ORDERS_XLSX); ws = wb['orders']; nid = _next_id_from_sheet(ORDERS_XLSX,'orders')
    ws.append([nid, order.get('customer_name'), order.get('address'), order.get('city'), order.get('phone'), json.dumps(order.get('items')), order.get('total_price'), order.get('created_at')]); wb.save(ORDERS_XLSX); wb.close()

def load_logs(limit=200):
    wb = load_workbook(LOGS_XLSX); ws = wb['logs']; logs=[]
    rows = list(ws.iter_rows(min_row=2, values_only=True)); rows = rows[-limit:][::-1]
    for row in rows: logs.append({'id':row[0],'product_id':row[1],'change_amount':row[2],'reason':row[3],'username':row[4],'timestamp':row[5]})
    wb.close(); return logs

def load_orders(limit=200):
    wb = load_workbook(ORDERS_XLSX); ws = wb['orders']; orders=[]
    rows = list(ws.iter_rows(min_row=2, values_only=True)); rows = rows[-limit:][::-1]
    for r in rows: orders.append({'id':r[0],'customer_name':r[1],'address':r[2],'city':r[3],'phone':r[4],'items':json.loads(r[5] or '[]'),'total_price':float(r[6] or 0),'created_at':r[7]})
    wb.close(); return orders

def load_users():
    with open(USERS_JSON,'r',encoding='utf-8') as f: return json.load(f)['users']
